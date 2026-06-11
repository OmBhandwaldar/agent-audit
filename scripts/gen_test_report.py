"""
Generate test-report.xlsx from a live pytest + coverage run.

Runs the suite, parses the JUnit results and the coverage JSON, and writes a
two-sheet workbook:
  - "Test Results": every test (file, name, status, duration)
  - "Coverage":     per-module statements / missed / percent, with the total

Usage:
  python scripts/gen_test_report.py            # writes ./test-report.xlsx
  python scripts/gen_test_report.py out.xlsx   # custom path
"""

import json
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
COV_PACKAGES = [
    "agentaudit", "api", "batcher", "crypto", "sdk",
    "tenancy", "agent", "ipfs", "algorand",
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PASS_FILL = PatternFill("solid", fgColor="DCFCE7")
FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")
TOTAL_FONT = Font(bold=True)


def run_pytest(junit_path: Path, cov_path: Path) -> int:
    """Run the suite, emitting JUnit XML and a coverage JSON. Returns pytest's exit code."""
    cov_args = [f"--cov={pkg}" for pkg in COV_PACKAGES]
    cmd = [
        sys.executable, "-m", "pytest", "-q",
        f"--junitxml={junit_path}",
        *cov_args,
        f"--cov-report=json:{cov_path}",
    ]
    return subprocess.run(cmd, cwd=ROOT).returncode


def parse_results(junit_path: Path) -> list[dict]:
    """Parse JUnit XML into [{file, name, status, seconds}] rows."""
    root = ET.parse(junit_path).getroot()
    rows = []
    for case in root.iter("testcase"):
        if case.find("failure") is not None or case.find("error") is not None:
            status = "FAIL"
        elif case.find("skipped") is not None:
            status = "SKIP"
        else:
            status = "PASS"
        rows.append({
            "file": case.get("classname", "").replace(".", "/"),
            "name": case.get("name", ""),
            "status": status,
            "seconds": round(float(case.get("time", 0.0)), 3),
        })
    return rows


def parse_coverage(cov_path: Path) -> tuple[list[dict], dict]:
    """Parse coverage JSON into per-file rows and the totals."""
    data = json.loads(cov_path.read_text())
    rows = []
    for path, info in sorted(data["files"].items()):
        s = info["summary"]
        rows.append({
            "module": path,
            "statements": s["num_statements"],
            "missed": s["missing_lines"],
            "percent": round(s["percent_covered"], 1),
        })
    t = data["totals"]
    totals = {
        "statements": t["num_statements"],
        "missed": t["missing_lines"],
        "percent": round(t["percent_covered"], 1),
    }
    return rows, totals


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 60)


def build_workbook(results: list[dict], cov_rows: list[dict], totals: dict) -> Workbook:
    """Assemble the two-sheet workbook."""
    wb = Workbook()

    # Sheet 1 — Test Results
    ws = wb.active
    ws.title = "Test Results"
    passed = sum(r["status"] == "PASS" for r in results)
    ws.append(["AgentAudit — Test Results"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"{passed}/{len(results)} passed", f"generated {datetime.now():%Y-%m-%d %H:%M}"])
    ws.append([])
    head_row = ws.max_row + 1
    ws.append(["File", "Test", "Status", "Duration (s)"])
    for r in results:
        ws.append([r["file"], r["name"], r["status"], r["seconds"]])
        fill = PASS_FILL if r["status"] == "PASS" else FAIL_FILL if r["status"] == "FAIL" else None
        if fill:
            ws.cell(row=ws.max_row, column=3).fill = fill
    # style the header row (row index head_row)
    for c in range(1, 5):
        cell = ws.cell(row=head_row, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    _autosize(ws)
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    # Sheet 2 — Coverage
    cs = wb.create_sheet("Coverage")
    cs.append(["Module", "Statements", "Missed", "Cover %"])
    _style_header(cs, 4)
    for r in cov_rows:
        cs.append([r["module"], r["statements"], r["missed"], r["percent"]])
    cs.append(["TOTAL", totals["statements"], totals["missed"], totals["percent"]])
    for c in range(1, 5):
        cs.cell(row=cs.max_row, column=c).font = TOTAL_FONT
    _autosize(cs)
    cs.freeze_panes = "A2"

    return wb


def main() -> None:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "test-report.xlsx"
    with tempfile.TemporaryDirectory() as tmp:
        junit_path = Path(tmp) / "junit.xml"
        cov_path = Path(tmp) / "cov.json"
        code = run_pytest(junit_path, cov_path)
        results = parse_results(junit_path)
        cov_rows, totals = parse_coverage(cov_path)

    wb = build_workbook(results, cov_rows, totals)
    wb.save(out)
    passed = sum(r["status"] == "PASS" for r in results)
    print(f"\nWrote {out}")
    print(f"  Test Results: {passed}/{len(results)} passed")
    print(f"  Coverage:     {totals['percent']}% across {len(cov_rows)} modules")
    sys.exit(code)


if __name__ == "__main__":
    main()
